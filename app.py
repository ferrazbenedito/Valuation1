import pandas as pd
import numpy as np
import scipy.stats as stats 
from scipy.stats import norm
import streamlit as st
pd.set_option("display.max_rows", None)
pd.set_option("display.max_columns", None)
pd.options.display.float_format = '{:.2f}'.format
import numpy_financial as npf
import psycopg2
from psycopg2 import sql
import urllib.parse as urlparse
import os


st.set_page_config(
    page_title="BJC - Real Estate",  # Title of the tab
    page_icon="🏘️",  # Icon for the tab
    layout="centered",  # Layout: "centered" or "wide"
    initial_sidebar_state="expanded",  # Sidebar state: "expanded", "collapsed", or "auto"
)

st.markdown(
    """
    <style>
    [data-testid="stSidebar"] {
        width: 500px; /* Adjust the width as needed */
    }
    </style>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:

    with st.columns(3)[1]:
        st.image("Gauss_Gap.jpg", width= 120, use_container_width=False)

    # Streamlit app
    st.title("Avaliação de Empreendimento Imobiliário")
    st.markdown("Coloque os valores para os cálculos de viabilidade do empreendimento.")

    with st.expander("Dados Básicos", expanded=True):
        Nome_da_Planilha = st.text_input("Nome da Planilha", value="Real_Estate_Calculations")
        Numero_de_Lotes = st.number_input("Numero de Lotes", min_value=1, value=500)
        Empreend_Terrenista = st.radio( "Empreendedor ou Terrenista", options=["Empreendedor", "Terrenista"])
        Receit_Lucro = st.radio( "Parceria no lucro ou na receita", options=["Receita", "Lucro"])

    with st.expander("Vendas e Parcelamento", expanded=False):

        Meses_Para_Vender = st.number_input("Meses Para Vender", min_value=1, value=15, help="Colocar o tempo máximo que acha que levará para vender todos os lotes")
        Media_de_Venda = st.number_input("Mês De Maior Venda", value=2.0, help="Com o mês de maior venda, podemos calcular a curva de vendas a partir dele")
        Desvio_Padrao_Venda = st.number_input("Desvio Padrao Venda", value=5.0, help="Quanto mais alto o desvio padrão, mais espalhado será o gráfico de vendas")
        Percentual_empreendedor = st.number_input("Percentual Empreendedor", min_value=0.0, max_value=1.0, value=0.7)
  
       


        Valor_Do_Lote = st.number_input("Valor do Lote", value=109227.47)
        Numero_Vezes_Entrada = st.number_input("Numero de Vezes Entrada", min_value=1, value=3, help="Número de vezes que o cliente pode parcelar a entrada")
        Percentuais_entrada_por_mes = []

        for i in range(Numero_Vezes_Entrada):
            percentual = st.number_input(f"Percentual Entrada Mês {i+1}", value=.01, key=f"percent_{i}")
            Percentuais_entrada_por_mes.append(percentual)



        Num_parcelas = st.number_input("Numero de Parcelas", min_value=1, value=50)
        Tipo_de_parcelameto = st.selectbox("Tipo de Parcelamento", options=[1, 2, 3], format_func=lambda x: "Price" if x == 1 else "Gradiente Juros Simples" if x == 2 else "Gradiente Juros Composto", help="Como o juros serão Calculados")

        Juros_ao_ano = st.number_input("Juros ao Ano", value=0.08)
        Periodicidade_do_reajuste = st.selectbox("Periodicidade do Reajuste", options=["Anual", "Mensal", "Semestral"])

    with st.expander("Comissão e Impostos", expanded=False):

       
        Parcelamento_Comissao = st.number_input("Parcelamento Comissao", min_value=1, value=3, help= "A comissão pode ser parcelada em até quantas vezes")
        Dict_Comissao = []

        for i in range(Parcelamento_Comissao):
            percentual_comissao = st.number_input(f"Percentual Comissão Por Mês {i+1}", value=.01, key=f"percentual_comissao{i}")
            Dict_Comissao.append(percentual_comissao)

        Dict_Comissao = dict(zip(range(1, Parcelamento_Comissao + 1), Dict_Comissao))

        Taxa_Adm = st.number_input("Taxa Adm", value=0.1)
        Imposto = st.number_input("Imposto", value=0.0673)

    with st.expander("Aprovações e Obras", expanded=False):

        Total_Obra = st.number_input("Total Obra", value=16766723.42, help="Valor total da obra")
        Obra_Pre_Lancamento = st.number_input("Obra Pre Lancamento", value=1000000.0, help="Do valor total da obra quanto será gasto antes do lançamento")
        Aprovacao = st.number_input("Aprovacao", value=1000000.0, help="Valor para projetos e aprovações que será adicionado ao custo da obra")

        Meses_Pre_Lancamento = st.number_input("Meses Pre Lançamento", min_value=1, value=28, help="Número de meses que começará o gasto pré obra e aprovação antes do ínicio das vendas")
        Media_Pre_Lancamento = st.number_input("Més De Maior Gasto Antes Do Lançamento", value=26.0, help="Segue a mesma lógica da média de vendas para a curva de gastos antes do lançamento")
        Std_Dev_Pre_Lancamento = st.number_input("Desvio Padrão Dos Gastos Pré Lançamento", value=15, help="Segue a mesma lógica do desvio padrão de vendas para a curva de gastos antes do lançamento")

        Meses_Pos_Lancamento = st.number_input("Meses Totais Para A Concluão Das Obras", min_value=1, value=20)
        Media_Pos_Lancamento = st.number_input("Mês De Maior Gasto Nas Obras", value=12.0, help="Segue a mesma lógica da média de vendas para a curva de gastos pós lançamento")
        Std_Dev_Pos_Lancamento = st.number_input("Desvio Padrão Dos Gastos Pós Lançamento ", value=4.0, help="Segue a mesma lógica do desvio padrão de vendas para a curva de gastos pós lançamento")

    with st.expander("Opcionais", expanded=False):

        Vai_reajustar_Parcelas = st.radio("Vai reajustar o preço do lote?", options=["Não", "Sim"])
        if Vai_reajustar_Parcelas == "Sim":
            percentual_Reajuste = st.number_input("Percentual Reajuste", value=0.05, help="Percentual de reajuste do valor do lote")
            periodicidade_reajuste = st.number_input(r"A cada quantos % de venda", value=.01 )

        # Fazer od sois abaixo

        Adicionar_inflacao = st.radio("Adicionar Inflação?", options=["Não", "Sim"])
        if Adicionar_inflacao == "Sim":
            inflacao = st.number_input("Inflação ao ano", value=0.05, step=.0001 , help="Percentual de inflação")

        Adicionar_titulo_do_governo = st.radio("Adicionar uma taxa concorrente anual?", options=["Não", "Sim"])
        if Adicionar_titulo_do_governo == "Sim":
            taxa_titulo = st.number_input("Taxa do título", value=0.05, step=.0001, help="Percentual de juros do título")



def curva_normal_mult_valor(valor_total, range_de_valores, media , devs_pad):
    """
    valor_total = valor a ser multiplicado 
    range_de_valores =  valor e não lista, dos meses que irao ser calculados (equivalente aos x)
    media = Média
    devs_pad = Desvio parão
    """

    lista_meses_venda = range(1, (range_de_valores + 1))

    valor = norm.pdf( lista_meses_venda,loc= media, scale= devs_pad)
    valor /= np.sum(valor)
    return(valor * valor_total)

vendas_por_mes = curva_normal_mult_valor(Numero_de_Lotes, Meses_Para_Vender, Media_de_Venda, Desvio_Padrao_Venda)

vendas_por_mes_reajustada = vendas_por_mes.copy()

if Vai_reajustar_Parcelas == "Sim":

    colunas_reajuste_vendas = ['Percentual Venda Mensal', 'Acumulado Vendas', 'Modular', 'Inteiros', 'Inteiros Multiplicados', 'Acumulado por Modular', 'Inteiros Mais Um', 'Acumulado Por Modular por Subida', 'Soma IntMult por Acum PorM']

    df_reajuste_vendas = pd.DataFrame(0, index=range(Meses_Para_Vender),   columns=colunas_reajuste_vendas)
    

    percentual_venda = vendas_por_mes/ sum(vendas_por_mes)

    percentual_venda =   pd.Series(percentual_venda)
    percentual_venda_acumulado = percentual_venda.cumsum()

    df_reajuste_vendas['Percentual Venda Mensal'] = percentual_venda
    df_reajuste_vendas['Acumulado Vendas'] = percentual_venda_acumulado
    periodicidade_reajuste = .1
    percentual_reajuste = .05

    df_reajuste_vendas['Modular'] = df_reajuste_vendas['Acumulado Vendas'] % periodicidade_reajuste
    df_reajuste_vendas['Inteiros'] = (df_reajuste_vendas['Acumulado Vendas'] // periodicidade_reajuste) -1
    df_reajuste_vendas['Inteiros Multiplicados'] = 1* (df_reajuste_vendas['Inteiros'] * percentual_reajuste) +1
    df_reajuste_vendas['Acumulado por Modular'] = df_reajuste_vendas['Modular'] / df_reajuste_vendas['Acumulado Vendas']	
    df_reajuste_vendas['Inteiros Mais Um'] = df_reajuste_vendas['Inteiros'] + 1
    df_reajuste_vendas['Acumulado Por Modular por Subida'] = df_reajuste_vendas['Acumulado por Modular'] * percentual_reajuste
    df_reajuste_vendas['Soma IntMult por Acum PorM'] = df_reajuste_vendas['Inteiros Multiplicados'] + df_reajuste_vendas['Acumulado Por Modular por Subida']

    lista_vendas_reajustadas = df_reajuste_vendas['Soma IntMult por Acum PorM'].tolist()
    vendas_por_mes_reajustada = lista_vendas_reajustadas * vendas_por_mes

















dict_vendas = dict(zip(range(1, (Meses_Para_Vender + 1)), vendas_por_mes_reajustada))

numero_vezes_entrada = len(Percentuais_entrada_por_mes)

valor_entrada_por_mes = [x * Valor_Do_Lote for x in Percentuais_entrada_por_mes ]
dict_entrada_por_mes = dict(zip(range(1, (numero_vezes_entrada +1)), valor_entrada_por_mes))

df_entrada = pd.DataFrame(np.zeros((Meses_Para_Vender , Meses_Para_Vender + Numero_Vezes_Entrada -1),dtype=int))
df_entrada.columns = [f"Mês {i+1}" for i in range(len(df_entrada.columns))]

def dicionario_para_index_df(df, dicionario, Mes_comeco):
    for i in range(len(df)):
        for j, key in enumerate(dicionario, start=1):  
            col_index = (j + i + Mes_comeco- 1) % len(df.columns)  
            df.iloc[i, col_index] = dicionario[key]

dicionario_para_index_df(df_entrada, dict_entrada_por_mes, 0)

adjusted_dict = {key - 1: value for key, value in dict_vendas.items()}

row_series = pd.Series(adjusted_dict)

df_resultado_entrada = df_entrada.mul(row_series, axis=0)

dict_entrada_final = df_resultado_entrada.sum().to_dict()

#Fazer a parcela na price
def juros_ano_mes(Taxa):
    Taxa =  ((Taxa+1) ** ((1/12))-1)
    return Taxa


Saldo_devedor = Valor_Do_Lote - sum(dict_entrada_por_mes.values())

Valor_parcela = round(-npf.pmt((juros_ano_mes(Juros_ao_ano)), Num_parcelas, Saldo_devedor), 2)


lista_fluxo_temp = list(range(len(Percentuais_entrada_por_mes) +1 , (len(Percentuais_entrada_por_mes) +1 + Num_parcelas )))



















if Tipo_de_parcelameto == 1:
    dict_parcelas = {chave: Valor_parcela for chave in lista_fluxo_temp}

elif Tipo_de_parcelameto == 2:
    lista_parcelas = list(range(0,   Num_parcelas))
    dict_parcelas = dict(zip(lista_fluxo_temp, [((1 + (Juros_ao_ano * int(i / 12))) * (Saldo_devedor / Num_parcelas)) for i in lista_parcelas]))

elif Tipo_de_parcelameto == 3:
    lista_parcelas = list(range(0,   Num_parcelas))
    dict_parcelas = dict(zip(lista_fluxo_temp, [(((1 + Juros_ao_ano) ** int(i / 12)) * (Saldo_devedor / Num_parcelas)) for i in lista_parcelas]))


if Adicionar_inflacao == "Sim":

    for key in dict_parcelas.keys():
        dict_parcelas[key] = (((1 + inflacao) ** ((key - Numero_Vezes_Entrada) // 12)) * dict_parcelas[key])








df_parcelas = pd.DataFrame((np.zeros((Meses_Para_Vender , Meses_Para_Vender + Numero_Vezes_Entrada -1 + Num_parcelas),dtype=int)))
df_parcelas.columns = [f"Mês {i+1}" for i in range(len(df_parcelas.columns))]

dicionario_para_index_df(df_parcelas, dict_parcelas, len(Percentuais_entrada_por_mes))



df_resultado_parcelas = df_parcelas.mul(row_series, axis=0)
dict_parcela_final = df_resultado_parcelas.sum().to_dict()
print(dict_parcela_final)
colunas_recebimento = ['Entrada', 'Parcela', 'Receita Bruta', 'Comissão (Não Incidência De Imposto)', 'Admin Carteira', 'Imposto', 'Outros Custos', 'Faturamento Líquido']

num_months = len(dict_parcela_final) +1
index = [f"Mês {i}" for i in range(0, num_months )]
df_recebimento = pd.DataFrame(0, columns=colunas_recebimento , index=index)

def inserir_dicionario(df, coluna, dicionario):
    for key, value in dicionario.items():
        if key in df.index:
            df.at[key, coluna] = value  
    return df

inserir_dicionario(df_recebimento, 'Entrada', dict_entrada_final)
inserir_dicionario(df_recebimento, 'Parcela', dict_parcela_final)
df_recebimento['Receita Bruta'] = df_recebimento['Entrada'] + df_recebimento['Parcela']

def Calcular_Valor_Comissão(dict_entrada, valor_do_lote, meses_para_vender, numero_vezes_entrada,dict_vendas, mes_comeco):

   
    dict_valor_entrada = {key: value * valor_do_lote for key, value in dict_entrada.items()}

    df_entrada_columns = [f"Mês {i}" for i in range(1, meses_para_vender + 1 + numero_vezes_entrada)]
    
    df_entrada = pd.DataFrame(0, columns=df_entrada_columns, index=range(1, meses_para_vender + 1))
    
    dicionario_para_index_df(df_entrada, dict_valor_entrada,0)

    for chave, valor in dict_vendas.items():
        for index, row in df_entrada.iterrows():
            if index == chave:
                df_entrada.loc[index] = valor * row  
    # Calculate the sum for each column and convert it to a dictionary
  

    soma_entrada_mes = dict(zip(range(1, len(df_entrada.columns) + 1), df_entrada.sum()))
    for key in soma_entrada_mes:
        soma_entrada_mes[key] *= -1

    return dict_valor_entrada, df_entrada, soma_entrada_mes

a, b, soma_comissao_mes = Calcular_Valor_Comissão(Dict_Comissao, Valor_Do_Lote, Meses_Para_Vender, len(Dict_Comissao),dict_vendas,3)

dict_comissao_final = (-b).sum().to_dict()

new_start_month = 1
dict_comissao_final = {f"Mês {i}": value for i, value in enumerate(dict_comissao_final.values(), start=new_start_month)}


inserir_dicionario(df_recebimento, 'Comissão (Não Incidência De Imposto)', dict_comissao_final)
df_recebimento['Admin Carteira'] = (df_recebimento['Receita Bruta'] + df_recebimento['Comissão (Não Incidência De Imposto)']) * -Taxa_Adm
df_recebimento['Imposto'] = (df_recebimento['Receita Bruta'] + df_recebimento['Comissão (Não Incidência De Imposto)']) * -Imposto
df_recebimento['Faturamento Líquido'] = df_recebimento['Receita Bruta'] + df_recebimento['Comissão (Não Incidência De Imposto)'] + df_recebimento['Admin Carteira'] + df_recebimento["Imposto"]


index_custo = [f"Mês {i}" for i in range(-Meses_Pre_Lancamento, num_months )]

df_custos = pd.DataFrame(0, columns=['Pré Obra','Pós Obra', 'Obra Total'], index=index_custo)


def curva_normal_mult_valor(valor_total, range_de_valores, media , devs_pad):
    """
    valor_total = valor a ser multiplicado 
    range_de_valores =  valor e não lista, dos meses que irao ser calculados (equivalente aos x)
    media = Média
    devs_pad = Desvio parão
    """

    lista_meses_venda = range(1, (range_de_valores + 1))

    valor = norm.pdf( lista_meses_venda,loc= media, scale= devs_pad)
    valor /= np.sum(valor)
    return(valor * valor_total)

custo_pre_Lancamento = Aprovacao +Obra_Pre_Lancamento
custo_pos_lancamento = Total_Obra - Obra_Pre_Lancamento

pre_obra_mes = curva_normal_mult_valor(custo_pre_Lancamento, Meses_Pre_Lancamento, Media_Pre_Lancamento, Std_Dev_Pre_Lancamento)
pos_obra_mes = curva_normal_mult_valor(custo_pos_lancamento, Meses_Pos_Lancamento, Media_Pos_Lancamento, Std_Dev_Pos_Lancamento )

dict_pre_obra = dict(zip([f"Mês {i}" for i in range(-Meses_Pre_Lancamento, 0 )], -pre_obra_mes ))
dict_pos_obra = dict(zip([f"Mês {i}" for i in range(0, Meses_Pos_Lancamento  )], -pos_obra_mes ))
inserir_dicionario(df_custos, 'Pré Obra', dict_pre_obra)  
inserir_dicionario(df_custos, 'Pós Obra', dict_pos_obra)  
df_custos['Obra Total'] = df_custos['Pré Obra'] + df_custos['Pós Obra']

aumento_range = pd.DataFrame(0, columns= colunas_recebimento  ,index = [f"Mês {i}" for i in range(-Meses_Pre_Lancamento, 0 )])

df_final= pd.concat([df_recebimento, df_custos], axis=1)
df_final = df_final.sort_index(key=lambda x: [int(i.split(' ')[1]) for i in x])
df_final =df_final.fillna(0)





df_final['Fluxo De Caixa Empreendedor'] = 0
df_final['Fluxo De Caixa Empreendedor Acumulado'] = 0
df_final['Fluxo De Caixa Terrenista'] = 0
df_final['Lucro Contábil'] = 0


if Receit_Lucro == "Receita":
    df_final['Fluxo De Caixa Empreendedor'] = (df_final['Faturamento Líquido'] * Percentual_empreendedor) + df_final['Obra Total']
    df_final['Fluxo De Caixa Terrenista'] = (df_final['Faturamento Líquido'] * (1 - Percentual_empreendedor))
    df_final['Fluxo De Caixa Empreendedor Acumulado'] = df_final['Fluxo De Caixa Empreendedor'].cumsum() 


elif Receit_Lucro == "Lucro":
  
    df_final['Fluxo De Caixa Empreendedor'] = (df_final['Faturamento Líquido'] ) + df_final['Obra Total']
    df_final['Fluxo De Caixa Empreendedor Acumulado'] = df_final['Fluxo De Caixa Empreendedor'].cumsum()
    # Find the index of the first positive value in column B
    # last_negative = (df_final['Fluxo De Caixa Empreendedor Acumulado'] > 0).cumsum() = 1
    found_first_positive = (df_final['Fluxo De Caixa Empreendedor Acumulado'] > 0).cumsum() > 1
    df_final['Lucro Contábil'] = df_final['Faturamento Líquido'] * found_first_positive
    df_final['Fluxo De Caixa Terrenista'] = (df_final['Lucro Contábil'] * (1 - Percentual_empreendedor))
    df_final['Fluxo De Caixa Empreendedor'] = df_final['Fluxo De Caixa Empreendedor'] - df_final['Fluxo De Caixa Terrenista']
    df_final["Fluxo De Caixa Empreendedor Acumulado"	] = df_final['Fluxo De Caixa Empreendedor Acumulado'] - df_final['Fluxo De Caixa Terrenista'] 


colunas_nao_negativas = ['Lucro Contábil', 'Faturamento Líquido', 'Fluxo De Caixa Terrenista']

for coluna in colunas_nao_negativas:
    if (df_final[coluna] < 0).any():
        st.write(f"The column {coluna} has a negative value.")

irr_empreendedor = npf.irr(df_final['Fluxo De Caixa Empreendedor'].values)
def juros_mes_ano(Taxa):
    Taxa = (Taxa+1)**(12) -1
    return Taxa

df_final['Fluxo De Caixa Empreendedor Acumulado'] = df_final['Fluxo De Caixa Empreendedor'].cumsum()
Ponto_negativo_empreendedor = min(df_final['Fluxo De Caixa Empreendedor Acumulado'])

variables_dict = {
    "Numero_de_Lotes": Numero_de_Lotes,
    "Meses_Para_Vender": Meses_Para_Vender,
    "Empreend_Terrenista": Empreend_Terrenista,
    "Receit_Lucro": Receit_Lucro,	
    "Media_de_Venda": Media_de_Venda,
    "Desvio_Padrao_Venda": Desvio_Padrao_Venda,
    "Percentual_empreendedor": Percentual_empreendedor,
    "Valor_Do_Lote": Valor_Do_Lote,
    "Numero_Vezes_Entrada": Numero_Vezes_Entrada,
    "Percentuais_entrada_por_mes": Percentuais_entrada_por_mes,
    "Num_parcelas": Num_parcelas,
    "Tipo_de_parcelameto": Tipo_de_parcelameto,
    "Juros_ao_ano": Juros_ao_ano,
    "Periodicidade_do_reajuste": Periodicidade_do_reajuste,
    # "Percentual_Comissao": Percentual_Comissao,
    "Parcelamento_Comissao": Parcelamento_Comissao,
    "Dict_Comissao": Dict_Comissao,
    "Taxa_Adm": Taxa_Adm,
    "Imposto": Imposto,
    "Total_Obra": Total_Obra,
    "Obra_Pre_Lancamento": Obra_Pre_Lancamento,
    "Aprovacao": Aprovacao,
    "Meses_Pre_Lancamento": Meses_Pre_Lancamento,
    "Media_Pre_Lancamento": Media_Pre_Lancamento,
    "Std_Dev_Pre_Lancamento": Std_Dev_Pre_Lancamento,
    "Meses_Pos_Lancamento": Meses_Pos_Lancamento,
    "Media_Pos_Lancamento": Media_Pos_Lancamento,
    "Std_Dev_Pos_Lancamento": Std_Dev_Pos_Lancamento,
}
df_variables = pd.DataFrame(list(variables_dict.items()), columns=['Variable', 'Value'])

# Save both to Excel
output_file = 'Final_6.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_final.to_excel(writer, sheet_name='DataFrame', index=False)
    df_variables.to_excel(writer, sheet_name='Variables', index=False)

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
razao_ponto_neg_Gasto_total = -(Ponto_negativo_empreendedor/(Total_Obra + Aprovacao))

irr_ano_temp = juros_mes_ano(irr_empreendedor)

additional_data = [
    ['IRR Ao Mês, Fluxo Geral', irr_empreendedor ],
    ['IRR Ao Ano, Fluxo Geral', irr_ano_temp],
    ['Ponto mínimo', Ponto_negativo_empreendedor],
    ['Ponto mais negativo do caixa dividido por gasto total', razao_ponto_neg_Gasto_total],
]
percentage_style = NamedStyle(name='percentage', number_format='0.00%')


with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_final.to_excel(writer, sheet_name='DataFrame', index=True)
    df_variables.to_excel(writer, sheet_name='Variables', index=False)


workbook = load_workbook(output_file)
sheet = workbook['DataFrame']

df_final_colunms = len(df_final.columns)

row_start = 3
for i, row in enumerate(additional_data, start=row_start):
    sheet.cell(row=i, column= df_final_colunms + 3, value=row[0])  # First column
    sheet.cell(row=i, column= df_final_colunms + 4, value=row[1])  # Second column


for sheet in workbook.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == irr_empreendedor:  # Check if the cell value matches the variable
                cell.style = percentage_style  # Apply percentage style

for sheet in workbook.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == irr_ano_temp:  # Check if the cell value matches the variable
                cell.style = percentage_style  # Apply percentage style

# Save the workbook
workbook.save(output_file)
vendas_por_mes_df = pd.DataFrame(vendas_por_mes, columns=['Lotes Por Mês'])
if "calculos_generated" not in st.session_state:
    st.session_state["calculos_generated"] = False

    

if st.button("Gerar Calculos"):
    # Set the flag to True when the button is pressed
    st.session_state["calculos_generated"] = True

acumulado_empreendedor = df_final["Fluxo De Caixa Empreendedor Acumulado"].tolist()

fluxo_terrenista = df_final["Fluxo De Caixa Terrenista"].tolist()
# Display calculations and toggle options only if the button was pressed
if st.session_state["calculos_generated"]:

    tab1, tab2, tab3, tab4 = st.tabs(["Empreendedor", "Terrenista", "Comparativo Indicadores", "Valores De Venda"])

    with tab1:

        col1, col2 = st.columns(2)  


        with col1:
            st.write('IRR Ao Mês, Fluxo Geral: {:.2%}'.format(irr_empreendedor).replace('.', ','))
            st.write('IRR Ao Ano, Fluxo Geral: {:.2%}'.format(irr_ano_temp).replace('.', ','))
            st.write('Ponto mínimo:  R$ {:,.2f}'.format(Ponto_negativo_empreendedor)
                    .replace(',', 'X').replace('.', ',').replace('X', '.'))
            st.write('Ponto mais negativo do caixa dividido por gasto total: {:.2%}'.format(razao_ponto_neg_Gasto_total).replace('.', ','))
        
        with col2:
            if Tipo_de_parcelameto == 1:
                st.write(f'Primeira Parcela: R$ {Valor_parcela:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'))

            else:
                st.write(f'Primeira Parcela: R$ {((Valor_Do_Lote- (sum(dict_entrada_por_mes.values())))/Num_parcelas):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'))
                

        # Create a toggle to switch between chart and DataFrame
        view_option = st.radio("Selecione a vizualização:", ("Gráfico", "Tabela"))

        

        if view_option == "Gráfico" and Empreend_Terrenista == "Empreendedor":
            st.subheader("Vendas por Mês")
            st.bar_chart(vendas_por_mes)
        elif view_option == "Tabela" and Empreend_Terrenista == "Empreendedor":
            st.subheader("DataFrame")
            st.dataframe(vendas_por_mes_df)
        elif view_option == "Gráfico" and Empreend_Terrenista == "Terrenista":
            st.subheader("Fluxo de Caixa Terrenista")
            st.line_chart(fluxo_terrenista)
        elif view_option == "Tabela" and Empreend_Terrenista == "Terrenista":
            st.subheader("DataFrame")
            st.dataframe(df_final["Fluxo De Caixa Terrenista"])
            st.write('Valor Recebido Terrenista Total R$ {:,.2f}'.format((df_final["Fluxo De Caixa Terrenista"].sum()))
                .replace(',', 'X').replace('.', ',').replace('X', '.'))
            st.write('VGV Terrenista R$ {:,.2f}'.format((Valor_Do_Lote * Numero_de_Lotes * (1 - Percentual_empreendedor)))
                .replace(',', 'X').replace('.', ',').replace('X', '.'))
        # st.bar_chart(acumulado_empreendedor)
#--------------------------------------------------------------------------------------------------------------



    with tab2:
        st.number_input("Qual o valor do área?", value=1000000)
        st.number_input("Quanto tempo do primero pagamento da área até o ínicio das vendas?", value=10)
        compra_da_area_parcelamento = st.radio("Forma da compra?", options=["A vista", "Parcelado (entrada + parcelas fixas)", "Costomizado"])

        lista_terrenista_recebimento = df_final["Fluxo De Caixa Terrenista"].tolist()
        lista_index_df_final = df_final.index.tolist()
        
        df_terrenista = {
            'Mês': lista_index_df_final,
            'Fluxo de Caixa Terrenista': lista_terrenista_recebimento,
        }
        df_terrenista = pd.DataFrame(df_terrenista)
        index_mes1 = df_terrenista[df_terrenista['Mês'] == 'Mês 0'].index[0]



        
        if compra_da_area_parcelamento == "A vista":
            valor_a_vista = st.number_input("Valor pago", value=10000)
            num_pareclas_antes_lancamento = st.number_input("Número de meses antes do lançamento", value=10)    
            # lista_parcelamento  = {}  
            lista_parcelamento = [-valor_a_vista]
            start_index = index_mes1 - num_pareclas_antes_lancamento
            start_index = max(start_index, 0)

            lista_parcelamento = dict(zip(range(start_index, start_index + len(lista_parcelamento)), lista_parcelamento))
            
        
        elif compra_da_area_parcelamento == "Parcelado (entrada + parcelas fixas)":
            valor_a_vista = st.number_input("Valor da entrada", value=10000)
            num_pareclas_antes_lancamento = st.number_input("Número de meses antes do lançamento", value=20)
            numero_parcelas_compra_terrenista = st.number_input("Quantas parcelas?", value=10)

            valor_parcela_fixa = st.number_input("Valor da parcela", value=100)
            
            lista_parcelamento = [-valor_a_vista] + [-valor_parcela_fixa] * numero_parcelas_compra_terrenista
            
            lista_terrenista_recebimento = df_final["Fluxo De Caixa Terrenista"].tolist()
            lista_index_df_final = df_final.index.tolist()
            
           
            start_index = index_mes1 - num_pareclas_antes_lancamento
            start_index = max(start_index, 0)
            
            lista_parcelamento = dict(zip(range(start_index, start_index + len(lista_parcelamento)), lista_parcelamento))
            
           
            

        elif compra_da_area_parcelamento == "Costomizado":

            valor_a_vista = st.number_input("Valor da entrada", value=10000)
            num_pareclas_antes_lancamento = st.number_input("Número de meses antes do lançamento", value=20)
            numero_parcelas_compra_terrenista = st.number_input("Quantas parcelas?", value=10)
            
            start_index = index_mes1 - num_pareclas_antes_lancamento
            start_index = max(start_index, 0)

            
            lista_mes_parcela = []
            lista_valor_parcela = []

            for i in range(numero_parcelas_compra_terrenista):
                col1, col2, col3, col4= st.columns(4)
                with col1:
                    mes_da_parcela = st.number_input(f"Meses pagamento {i+1}", value=(5 + i), key=f"Pagamento {i}") 
                    lista_mes_parcela.append(mes_da_parcela)

                with col3:
                    valor_parcela_compra_area = st.number_input(f"Valor da parcela {i+1}", value=1000, key=f"Valor {i}") *-1
                    lista_valor_parcela.append(valor_parcela_compra_area)
            lista_parcelamento = {}
            

            lista_parcelamento = dict(zip(lista_mes_parcela, lista_valor_parcela))
            lista_parcelamento = {key + start_index: value for key, value in lista_parcelamento.items()}
            lista_parcelamento[start_index] =  -valor_a_vista

            # st.write(lista_parcelamento)

#--------------------------------------------------------------------------------------------------------------

        lista_terrenista_recebimento = df_final["Fluxo De Caixa Terrenista"].tolist()
        lista_index_df_final = df_final.index.tolist()
        
        df_terrenista = {
            'Mês': lista_index_df_final,
            'Fluxo de Caixa Terrenista': lista_terrenista_recebimento,
        }
        df_terrenista = pd.DataFrame(df_terrenista)
        # df_terrenista['Fluxo Terrenista Geral'] = df_terrenista.set_index('Mês')
        # st.write(lista_index_df_final)
        index_mes1 = df_terrenista[df_terrenista['Mês'] == 'Mês 0'].index[0]
        
        # Calculate the starting index for inserting the list
        start_index = index_mes1 - num_pareclas_antes_lancamento
        

        # Ensure the index does not go below 0
        start_index = max(start_index, 0)
        
        # lista_parcelamento = dict(zip(range(start_index, start_index + len(lista_parcelamento)), lista_parcelamento))
        
        # Add a new column 'C' filled with 0
        df_terrenista['Fluxo Terrenista Geral'] = 0
        inserir_dicionario(df_terrenista, 'Fluxo Terrenista Geral', lista_parcelamento)
        df_terrenista['Fluxo Terrenista Geral'] = df_terrenista['Fluxo Terrenista Geral'] + df_terrenista['Fluxo de Caixa Terrenista']
        irr_terrenista = npf.irr(df_terrenista['Fluxo Terrenista Geral'].values)
        irr_ano_temp_terrenista = juros_mes_ano(irr_terrenista) 
        st.write('IRR Ao Mês, Terrenista: {:.2%}'.format(irr_terrenista).replace('.', ','))
        st.write('IRR Ao Ano, Terrenista: {:.2%}'.format(irr_ano_temp_terrenista).replace('.', ','))
        st.dataframe(df_terrenista)

    

    with tab3:
        if Adicionar_inflacao == "Sim" and Adicionar_titulo_do_governo == "Sim":
            
            # Company performance: annual IRR computed from your cash flows
            company_annual_irr = irr_ano_temp

            # Alternative investment: 3% plus inflation.
            # For example, if inflation is assumed to be 2%, the annual rate is:
            alternative_annual_rate = inflacao + taxa_titulo  # 5% annual nominal

            # Determine the period for the comparison.
            # Your cash flows span 19 months (the example list has 19 values).
            n_months = len(df_final['Fluxo De Caixa Empreendedor'])
            months = np.arange(0, n_months + 1)  # from month 0 to month 19

            # -----------------------------------------------------
            # Convert annual rates to equivalent monthly growth rates
            # -----------------------------------------------------

            # For compound growth, the monthly rate is:
            company_monthly_rate = (1 + company_annual_irr) ** (1/12) - 1
            alternative_monthly_rate = (1 + alternative_annual_rate) ** (1/12) - 1

            # -----------------------------------------------------
            # Create time series for a $1 investment growing over time
            # -----------------------------------------------------

            # Using compound interest formulas:
            company_growth = (1 + company_monthly_rate) ** months
            alternative_growth = (1 + alternative_monthly_rate) ** months

            # -----------------------------------------------------
            # Create a DataFrame for the two series
            # -----------------------------------------------------

            data = pd.DataFrame({
                'Mês': months,
                'Retorno Empreendimento': company_growth,
                'Retorno Investimento Alternativo': alternative_growth
            })
            data.set_index('Mês', inplace=True)

            # -----------------------------------------------------
            # Streamlit app display
            # -----------------------------------------------------

            st.title("IRR Do Empreendiemneto vs. Investimento Alternativo")
            st.write(
                f"Retorno do empreendimento {company_annual_irr*100:.1f}% anual "
                f"comparado a um retorno de investimento alternativo de {alternative_annual_rate*100:.1f}% anual."
            )
            st.line_chart(data)


    with tab4:

        chart_valor_venda = st.radio("Selecione a vizualização:", ("Valor Do Lote", "Valor Entrada", "Valor Primeira Parcela"))
        if chart_valor_venda == "Valor Do Lote":
            st.bar_chart((vendas_por_mes_reajustada / vendas_por_mes) * Valor_Do_Lote)
        elif chart_valor_venda == "Valor Entrada":
            st.bar_chart((vendas_por_mes_reajustada / vendas_por_mes) * Valor_Do_Lote * sum(Percentuais_entrada_por_mes))
        elif chart_valor_venda == "Valor Primeira Parcela":
            st.bar_chart((vendas_por_mes_reajustada / vendas_por_mes) * Valor_parcela)








        
    try:
        output_file = 'Final_6.xlsx'  # Ensure this matches your generated file path
        with open(output_file, "rb") as file:
            st.download_button(
                label="Baixar Arquivo Excel",
                data=file,
                file_name="{}.xlsx".format(Nome_da_Planilha),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        
            
    except Exception as e:
        st.error(f"Erro ao gerar arquivo: {e}")


def create_input_changes_table():
    DATABASE_URL = os.getenv("DATABASE_URL")
    if not DATABASE_URL:
        st.error("DATABASE_URL is missing. Make sure you set it in Heroku.")
        return
    
    try:
        result = urlparse.urlparse(DATABASE_URL)
        db_host = result.hostname
        db_user = result.username
        db_password = result.password
        db_name = result.path.lstrip("/")
        db_port = result.port if result.port else 5432

        conn = psycopg2.connect(
            host=db_host,
            user=db_user,
            password=db_password,
            dbname=db_name,
            port=db_port,
        )
        conn.autocommit = True
        cur = conn.cursor()

        create_table_query = """
        CREATE TABLE IF NOT EXISTS input_changes (
            id SERIAL PRIMARY KEY,
            input_name TEXT,
            old_value TEXT,
            new_value TEXT,
            changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
        cur.execute(create_table_query)

        cur.close()
        conn.close()
    except psycopg2.Error as e:
        st.error(f"Database error: {e}")

def track_input_changes():
    if "previous_inputs" not in st.session_state:
        st.session_state["previous_inputs"] = {}

    current_inputs = {
        "Nome_da_Planilha": Nome_da_Planilha,
        "Numero_de_Lotes": Numero_de_Lotes,
        "Empreend_Terrenista": Empreend_Terrenista,
        "Receit_Lucro": Receit_Lucro,
        "Meses_Para_Vender": Meses_Para_Vender,
        "Media_de_Venda": Media_de_Venda,
        "Desvio_Padrao_Venda": Desvio_Padrao_Venda,
        "Percentual_empreendedor": Percentual_empreendedor,
        "Valor_Do_Lote": Valor_Do_Lote,
        "Numero_Vezes_Entrada": Numero_Vezes_Entrada,
        "Percentuais_entrada_por_mes": Percentuais_entrada_por_mes,
        "Num_parcelas": Num_parcelas,
        "Tipo_de_parcelameto": Tipo_de_parcelameto,
        "Juros_ao_ano": Juros_ao_ano,
        "Periodicidade_do_reajuste": Periodicidade_do_reajuste,
        "Parcelamento_Comissao": Parcelamento_Comissao,
        "Dict_Comissao": Dict_Comissao,
        "Taxa_Adm": Taxa_Adm,
        "Imposto": Imposto,
        "Total_Obra": Total_Obra,
        "Obra_Pre_Lancamento": Obra_Pre_Lancamento,
        "Aprovacao": Aprovacao,
        "Meses_Pre_Lancamento": Meses_Pre_Lancamento,
        "Media_Pre_Lancamento": Media_Pre_Lancamento,
        "Std_Dev_Pre_Lancamento": Std_Dev_Pre_Lancamento,
        "Meses_Pos_Lancamento": Meses_Pos_Lancamento,
        "Media_Pos_Lancamento": Media_Pos_Lancamento,
        "Std_Dev_Pos_Lancamento": Std_Dev_Pos_Lancamento,
    }
    if st.session_state["calculos_generated"]:
        for key, value in current_inputs.items():
            if key in st.session_state["previous_inputs"] and st.session_state["previous_inputs"][key] != value:
                save_input_change(key, st.session_state["previous_inputs"][key], value)
            st.session_state["previous_inputs"][key] = value


def save_input_change(input_name, old_value, new_value):
    DATABASE_URL = os.getenv("DATABASE_URL")
    if not DATABASE_URL:
        st.error("DATABASE_URL is missing. Make sure you set it in Heroku.")
        return
    
    try:
        result = urlparse.urlparse(DATABASE_URL)
        db_host = result.hostname
        db_user = result.username
        db_password = result.password
        db_name = result.path.lstrip("/")
        db_port = result.port if result.port else 5432

        conn = psycopg2.connect(
            host=db_host,
            user=db_user,
            password=db_password,
            dbname=db_name,
            port=db_port,
        )
        conn.autocommit = True
        cur = conn.cursor()

        insert_query = """
            INSERT INTO input_changes (input_name, old_value, new_value)
            VALUES (%s, %s, %s);
        """
        cur.execute(insert_query, (input_name, str(old_value), str(new_value)))

        cur.close()
        conn.close()
    except psycopg2.Error as e:
        st.error(f"Database error: {e}")

# Ensure the input_changes table exists
create_input_changes_table()

# Your existing code for inputs...

# Track and save input changes
track_input_changes()

